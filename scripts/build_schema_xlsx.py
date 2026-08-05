"""Generate Scryfall vs MTGJSON schema comparison workbook.

Single "Sources" sheet + "Legend" sheet.

Group is the anatomy of the data (Identity / Card text / Picture / Prices /
Legalities / Sets / …), NOT the source. Cross-source rows for the same
anatomical concept sit side-by-side so users can merge/delete them while
annotating decisions in the Status column.

Run from repo root: python scripts/build_schema_xlsx.py
Output: docs/schema-scryfall-vs-mtgjson.xlsx
"""
from __future__ import annotations

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COUNTS_PATH = "docs/schema-counts.json"
ROWS_PATH = "docs/schema-rows.json"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
OUR_FILL = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
SKIP_FILL = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
NEED_DECISION_FILL = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")

# (group, element, scryfall source, mtgjson source, app-today source, why)
# Groups are anatomical buckets — duplicate concepts from each source sit side-by-side.
APP_ELEMENTS = [

    # =========================================================================
    # IDENTITY — names, ids, collector number, layout
    # =========================================================================
    ("Identity", "Primary UUID", "Card.id", "Card.uuid", "cards.id (uuid)", "primary key — uuid comes from Scryfall, mirrored by MTGJSON"),
    ("Identity", "Oracle UUID", "Card.oracle_id", "Card.scryfallOracleId", "cards.oracle_id (uuid)", "groups all reprints of the same rules text"),
    ("Identity", "Card name", "Card.name", "Card.name", "cards.name (text)", ""),
    ("Identity", "ASCII name", "—", "Card.asciiName", "not stored", "ASCII transliteration for non-English card names"),
    ("Identity", "Layout", "Card.layout", "Card.layout", "not stored (per-face matters for rendering)", "normal / transform / modal_dfc / split / flip / adventure / …"),
    ("Identity", "Collector number", "Card.collector_number", "Card.number", "cards.collector_number (text)", "set-relative ('184a', '184★', 'TG01')"),
    ("Identity", "Sort key", "derived from Card.collector_number", "Card.numberSort (recommended)", "derived: hand-rolled sortKey() regex on collector_number", "MTGJSON's numberSort is sorted-friendly — would replace sortKey()"),
    ("Identity", "Language", "Card.lang", "Card.language", "filtered to 'en' on import", "ISO 639-1 — Scryfall default English only"),
    ("Identity", "Card release date", "Card.released_at", "Card.releaseDate", "not stored on cards (kept on sets)", "Card-level release date (sets.release_date aggregates this)"),

    # =========================================================================
    # CARD TEXT — oracle, flavor, printed, foreign
    # =========================================================================
    ("Card text", "Oracle text (card-level)", "Card.oracle_text", "Card.text (when not in faces)", "covered by cards.attrs.oracle_text", "rules text (English)"),
    ("Card text", "Flavor text (card-level)", "Card.flavor_text", "Card.flavorText", "covered by cards.attrs.flavor_text", "flavor / quote text"),
    ("Card text", "Printed name (foreign/localised)", "Card.printed_name", "Card.faceName (related)", "not stored", "printed name on the card (may differ from Card.name in foreign releases)"),
    ("Card text", "Printed text (foreign/localised)", "Card.printed_text", "Card.cardFaces[].text (when localised)", "not stored", "printed text on the card"),
    ("Card text", "Printed type line (foreign/localised)", "Card.printed_type_line", "Card.cardFaces[].type (when localised)", "not stored", "printed type line"),
    ("Card text", "Foreign / localised data", "—", "Card.foreignData[].{language,name,faceName,flavorText,text}", "not stored", "localised data — no demand; skip until multilingual UI lands"),
    ("Card text", "Rulings (date)", "Card.rulings[].published_at", "Card.rulings[].date", "not imported", "when the ruling was published"),
    ("Card text", "Rulings (text)", "Card.rulings[].comment", "Card.rulings[].text", "not imported", "ruling text"),

    # =========================================================================
    # PICTURE — image_uris, illustration_id, art_crop, border_crop
    # =========================================================================
    ("Picture", "Thumbnail", "Card.image_uris.small", "identifiers.imageSmallUrl (5e MB only)", "cards.image_small (text)", ""),
    ("Picture", "Default size", "Card.image_uris.normal", "—", "not stored (fallback to large)", "default size"),
    ("Picture", "Large", "Card.image_uris.large", "identifiers.imageLargeUrl (5e MB only)", "cards.image_large (text)", "used until 2026-08-04 for binder cover marquee"),
    ("Picture", "PNG (transparent)", "Card.image_uris.png", "—", "not stored", "lossless PNG with transparent background"),
    ("Picture", "Art crop", "Card.image_uris.art_crop", "—", "cards.image_art_crop (text)", "illustration only — no frame, no text"),
    ("Picture", "Border crop", "Card.image_uris.border_crop", "—", "not stored", "full card image with black border trimmed off"),
    ("Picture", "Illustration id", "Card.illustration_id", "Card.scryfallIllustrationId", "not stored", "stable per-artwork (reused across reprints) — would let us dedupe art across reprints"),

    # =========================================================================
    # GAMEPLAY — mana_cost, cmc, type_line, power/toughness/loyalty, keywords
    # =========================================================================
    ("Gameplay", "Mana cost", "Card.mana_cost", "Card.manaCost", "covered by cards.attrs.mana_cost", "{2}{W}{U} notation — tokenised per face in cardToMockFaces"),
    ("Gameplay", "Converted mana cost", "Card.cmc", "Card.manaValue", "covered by cards.attrs.cmc", ""),
    ("Gameplay", "Type line", "Card.type_line", "Card.type", "covered by cards.attrs.type_line", "includes supertypes + types + subtypes"),
    ("Gameplay", "Power", "Card.power", "Card.power", "covered by cards.attrs.power", "creature power ('*', '1+1', etc.)"),
    ("Gameplay", "Toughness", "Card.toughness", "Card.toughness", "covered by cards.attrs.toughness", "creature toughness"),
    ("Gameplay", "Loyalty", "Card.loyalty", "Card.loyalty", "covered by cards.attrs.loyalty", "planeswalker starting loyalty"),
    ("Gameplay", "Defense (battle)", "Card.card_faces[].defense", "Card.defense", "not stored", "battle card defense"),
    ("Gameplay", "Life modifier (Vanguard)", "Card.life_modifier", "Card.life", "not stored", "Vanguard-only"),
    ("Gameplay", "Hand modifier (Vanguard)", "Card.hand_modifier", "Card.hand", "not stored", "Vanguard-only"),
    ("Gameplay", "Keywords", "Card.keywords", "Card.keywords", "covered by cards.attrs.keywords", "keyword abilities (Flying, Trample, …)"),
    ("Gameplay", "Produced mana", "Card.produced_mana", "—", "not stored", "mana a permanent could produce (rare)"),
    ("Gameplay", "EDHREC rank", "—", "Card.edhrecRank", "not stored", "EDHREC popularity rank (0-9999+) — great for 'most-played Commander cards' sort"),
    ("Gameplay", "Per-face CMC", "Card.card_faces[].cmc", "Card.faceConvertedManaCost", "not stored", "per-face converted mana cost"),

    # =========================================================================
    # FACES — split/transform/MDFC/adventure: per-face fields
    # =========================================================================
    ("Faces", "Face container", "Card.card_faces[]", "Card.cardFaces[]", "covered by cards.attrs.card_faces[]", "per-face fields for transform/MDFC/split/adventure"),
    ("Faces", "Face 1 UUID", "Card.face_one_id", "—", "not stored", "Scryfall ID of front face — MTGJSON exposes otherFaceIds[] instead"),
    ("Faces", "Face 2 UUID", "Card.face_two_id", "—", "not stored", "back-face Scryfall ID"),
    ("Faces", "Per-face name", "Card.card_faces[].name", "Card.cardFaces[].name / Card.faceName", "covered by cards.attrs.card_faces[].name", ""),
    ("Faces", "Per-face flavor name", "Card.card_faces[].flavor_name", "Card.faceFlavorName", "not stored", "per-face flavour name (silver-bordered)"),
    ("Faces", "Per-face mana cost", "Card.card_faces[].mana_cost", "Card.cardFaces[].manaCost", "covered by cards.attrs.card_faces[].mana_cost", ""),
    ("Faces", "Per-face type line", "Card.card_faces[].type_line", "Card.cardFaces[].type", "covered by cards.attrs.card_faces[].type_line", ""),
    ("Faces", "Per-face oracle text", "Card.card_faces[].oracle_text", "Card.cardFaces[].text", "covered by cards.attrs.card_faces[].oracle_text", ""),
    ("Faces", "Per-face power/toughness/loyalty", "Card.card_faces[].{power,toughness,loyalty}", "Card.cardFaces[].{power,toughness,loyalty}", "covered by cards.attrs.card_faces[].*", ""),
    ("Faces", "Per-face flavor text", "Card.card_faces[].flavor_text", "Card.cardFaces[].flavorText", "covered by cards.attrs.card_faces[].flavor_text", ""),
    ("Faces", "Per-face color indicator", "Card.card_faces[].color_indicator", "Card.cardFaces[].colorIndicator / Card.colorIndicator", "not stored", "colour indicator dot for colourless cards with coloured mana cost (e.g. devoid)"),
    ("Faces", "Per-face image_uris", "Card.card_faces[].image_uris", "Card.cardFaces[].identifiers.imageUrls{...}", "covered by per-face image_art_crop fallback", "same five variants as card-level image_uris"),
    ("Faces", "Per-face defense", "Card.card_faces[].defense", "Card.cardFaces[].defense", "not stored", "battle card defense"),
    ("Faces", "Other face ids", "—", "Card.otherFaceIds", "not stored", "sister-face IDs across the same card"),
    ("Faces", "Side (split)", "Card.side", "Card.side", "not stored", "a / b for split cards"),

    # =========================================================================
    # COLORS & COLOUR IDENTITY — WUBRG, colorIdentity, indicator
    # =========================================================================
    ("Colors & colour identity", "Card colours", "Card.colors", "Card.colors", "cards.card_facets rows (derived)", "mana colours (WUBRG)"),
    ("Colors & colour identity", "Colour identity (Commander)", "Card.colors (includes colour indicator)", "Card.colorIdentity", "cards.card_facets rows (derived)", "Scryfall's 'colors' already covers this for our needs"),
    ("Colors & colour identity", "Colour indicator", "Card.card_faces[].color_indicator", "Card.colorIndicator", "not stored", "colour indicator dots"),
    ("Colors & colour identity", "WUBRG order", "—", "—", "hardcoded 'WUBRG' in util.ts and ComboSlicer.tsx", "standard MTG order — neither source enumerates it"),

    # =========================================================================
    # TYPES & TRIBES — parsed supertype/type/subtype arrays
    # =========================================================================
    ("Types & tribes", "Types[]", "derived from Card.type_line", "Card.types", "not stored", "supertypes parsed: ['Legendary','Creature']"),
    ("Types & tribes", "Subtypes[]", "derived from Card.type_line", "Card.subtypes", "not stored", "['Human','Wizard'] — would let us slice by tribe (Elves, Goblins, …) without substring hacks"),
    ("Types & tribes", "Supertypes[]", "derived from Card.type_line", "Card.supertypes", "not stored", "['Legendary','Basic','Snow']"),

    # =========================================================================
    # FRAME & FINISH — frame, security_stamp, border_color, finishes
    # =========================================================================
    ("Frame & finish", "Frame", "Card.frame", "Card.frameVersion", "covered by cards.attrs.frame", "1993 / 1997 / 2003 / 2015 / future"),
    ("Frame & finish", "Security stamp", "Card.security_stamp", "Card.securityStamp", "covered by cards.attrs.security_stamp", "oval / triangle / acorn / arena — primary UB signal"),
    ("Frame & finish", "Border colour", "Card.border_color", "Card.borderColor", "covered by cards.attrs.border_color", "black / white / borderless / silver / gold"),
    ("Frame & finish", "Watermark", "Card.watermark", "Card.watermark", "not stored", "card watermark — would be nice for set icon hover"),
    ("Frame & finish", "Finishes", "Card.finishes", "Card.finishes", "cards.finishes (text[])", "default ['nonfoil'] if missing"),
    ("Frame & finish", "Has foil", "derived from Card.finishes", "Card.hasFoil", "not stored (Card.finishes covers it)", "available in foil"),
    ("Frame & finish", "Has nonfoil", "derived from Card.finishes", "Card.hasNonFoil", "not stored (Card.finishes covers it)", "available in nonfoil"),

    # =========================================================================
    # PRINT — set, set_id, set_name, set_type, collector_number, release
    # =========================================================================
    ("Print", "Set code", "Card.set", "Card.setCode", "derived via set_id -> sets.code", "Scryfall lowercase; MTGJSON uppercase — we lowercase on import"),
    ("Print", "Set UUID", "Card.set_id", "Set.uuid (via /sets cross-ref)", "cards.set_id (uuid)", "stable set uuid"),
    ("Print", "Set name", "Card.set_name", "Set.name", "derived via set_id -> sets.name", "set display name"),
    ("Print", "Set type", "Card.set_type", "Set.type", "derived via set_id -> sets.set_type, bucketed by MTG_BUCKETS", "30+ enum values"),
    ("Print", "Set URI (Scryfall)", "Card.set_uri", "—", "not stored", "Scryfall API for the parent set object"),
    ("Print", "Set search URI", "Card.set_search_uri", "—", "not stored", "search URL for all cards in the set"),
    ("Print", "Scryfall set web URL", "Card.scryfall_set_uri", "—", "not stored", "web URL for the set"),
    ("Print", "Rulings URI", "Card.rulings_uri", "— (rulings already in Card.rulings[])", "covered by Rulings rows (not imported)", "rulings JSON for this card"),
    ("Print", "Prints search URI", "Card.prints_search_uri", "—", "not stored", "search URL for all reprints"),
    ("Print", "Collector number", "Card.collector_number", "Card.number", "covered by cards.collector_number (also in Identity group)", "duplicate — same field, second row kept so cross-source rows align for merge"),
    ("Print", "Story spotlight", "Card.story_spotlight", "Card.isStorySpotlight", "not stored", "featured in the weekly Story Spotlight"),
    ("Print", "Time-shifted", "—", "Card.isTimeshifted", "not stored", "Time Spiral timeshifted cards"),
    ("Print", "Oversized", "—", "Card.isOversized", "not stored", "Vanguard / planar cards"),

    # =========================================================================
    # PROMO & CONTENT — promo_types, booster, variation, content_warning
    # =========================================================================
    ("Promo & content", "Promo types", "Card.promo_types", "Card.promoTypes", "covered by cards.attrs.promo_types", "['universesbeyond','boosterfun','promo',…] — crossover signal"),
    ("Promo & content", "Booster (draft)", "Card.booster", "Card.isStarter (related)", "not stored", "available in draft boosters"),
    ("Promo & content", "Variation flag", "Card.variation", "—", "not stored", "is a variation of another card"),
    ("Promo & content", "Variation of (UUID)", "Card.variation_of", "—", "not stored", "the base card this varies from"),
    ("Promo & content", "Content warning", "Card.content_warning", "—", "not stored", "has sensitive content"),
    ("Promo & content", "Starter card", "derived (Card.booster-related)", "Card.isStarter", "not stored", "starter set card (e.g. Planeswalker Starter Kit)"),

    # =========================================================================
    # IDENTIFIERS (CROSS-SOURCE) — third-party product IDs (tcgplayer, mcm, arena…)
    # =========================================================================
    ("Identifiers (cross-source)", "TCGplayer product id", "Card.tcgplayer_id", "identifiers.tcgplayerProductId", "not stored", "TCGplayer product ID"),
    ("Identifiers (cross-source)", "TCGplayer etched id", "Card.tcgplayer_etched_id", "—", "not stored", "etched foil TCGplayer ID"),
    ("Identifiers (cross-source)", "Cardmarket id", "Card.cardmarket_id", "identifiers.mcmId", "not stored", "Cardmarket product ID (EU)"),
    ("Identifiers (cross-source)", "Cardmarket meta id", "—", "identifiers.mcmMetaId", "not stored", "Cardmarket meta ID"),
    ("Identifiers (cross-source)", "MTGO id", "Card.mtgo_id", "identifiers.mtgoId", "not stored", "MTGO catalogue ID"),
    ("Identifiers (cross-source)", "MTGO foil id", "Card.mtgo_foil_id", "— (deprecated in MTGJSON)", "not stored", "MTGO foil catalogue ID"),
    ("Identifiers (cross-source)", "Arena id", "Card.arena_id", "identifiers.mtgArenaId", "not stored", "MTG Arena ID"),
    ("Identifiers (cross-source)", "Multiverse id", "Card.multiverse_ids", "identifiers.multiverseId", "not stored", "legacy Gatherer IDs; deprecated"),
    ("Identifiers (cross-source)", "Card Kingdom id", "—", "identifiers.cardKingdomId", "not stored", "Card Kingdom product ID"),
    ("Identifiers (cross-source)", "Cardsphere id", "—", "identifiers.cardsphereId", "not stored", "Cardsphere ID"),
    ("Identifiers (cross-source)", "Artist ids", "—", "Card.artistIds", "not stored", "Wizards artist IDs — would let us group by artist across sets"),

    # =========================================================================
    # PRICES — usd, usd_foil, usd_etched, eur, tix
    # =========================================================================
    ("Prices", "Prices table", "Card.prices.{usd,usd_foil,usd_etched}", "Card.prices.paper.{tcgplayer,cardmarket}", "prices rows joined on card_id + finish + as_of", ""),
    ("Prices", "as_of", "derived: stamp at import run", "derived: stamp at import run", "prices.as_of (date)", "used by /value sparkline gates"),
    ("Prices", "USD (nonfoil)", "Card.prices.usd", "Card.prices.paper.tcgplayer.{normal,holofoil}", "covered by prices table, finish='nonfoil'", "TCGplayer USD market price (nonfoil)"),
    ("Prices", "USD (foil)", "Card.prices.usd_foil", "Card.prices.paper.tcgplayer.foil", "covered by prices table, finish='foil'", "foil price"),
    ("Prices", "USD (etched)", "Card.prices.usd_etched", "— (MTGJSON's hasEtched isn't priced)", "covered by prices table, finish='etched'", "etched foil price"),
    ("Prices", "EUR", "Card.prices.eur", "Card.prices.paper.cardmarket", "not stored", "Cardmarket EUR"),
    ("Prices", "EUR (foil)", "Card.prices.eur_foil", "Card.prices.paper.cardmarket.foil", "not stored", "Cardmarket EUR foil"),
    ("Prices", "TIX", "Card.prices.tix", "— (MTGO is shut)", "not stored", "MTGO tickets"),

    # =========================================================================
    # LEGALITIES — all 17 formats
    # =========================================================================
    ("Legalities", "Card legalities object", "Card.legalities (17 keys)", "Card.legalities (17 keys)", "covered by cards.attrs.legalities, aggregated into sets.legalities", "Legal / Not Legal / Banned / Restricted"),
    ("Legalities", "Standard", "Card.legalities.standard", "Card.legalities.standard", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Pioneer", "Card.legalities.pioneer", "Card.legalities.pioneer", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Modern", "Card.legalities.modern", "Card.legalities.modern", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Legacy", "Card.legalities.legacy", "Card.legalities.legacy", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Vintage", "Card.legalities.vintage", "Card.legalities.vintage", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Commander", "Card.legalities.commander", "Card.legalities.commander", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Pauper", "Card.legalities.pauper", "Card.legalities.pauper", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Oathbreaker", "Card.legalities.oathbreaker", "Card.legalities.oathbreaker", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Brawl", "Card.legalities.brawl", "Card.legalities.brawl", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Historic", "Card.legalities.historic", "Card.legalities.historic", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Alchemy", "Card.legalities.alchemy", "Card.legalities.alchemy", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Explorer", "Card.legalities.explorer", "Card.legalities.explorer", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Gladiator", "Card.legalities.gladiator", "Card.legalities.gladiator", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Standardbrawl", "Card.legalities.standardbrawl", "Card.legalities.standardbrawl", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Timeless", "Card.legalities.timeless", "Card.legalities.timeless", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Predh", "Card.legalities.predh", "Card.legalities.predh", "covered by cards.attrs.legalities", ""),
    ("Legalities", "Premodern", "Card.legalities.premodern", "Card.legalities.premodern", "covered by cards.attrs.legalities", ""),

    # =========================================================================
    # RELATED CARDS — all_parts, preview
    # =========================================================================
    ("Related cards", "All parts (combo/meld)", "Card.all_parts[]", "Card.parts / Card.tokenParts / Card.signature", "not stored", "combo pieces, meld results, meld inputs"),
    ("Related cards", "Preview date", "Card.preview.previewed_at", "—", "not stored", "first preview date"),
    ("Related cards", "Preview source URI", "Card.preview.source_uri", "—", "not stored", "preview URL"),
    ("Related cards", "Preview source name", "Card.preview.source", "—", "not stored", "preview source name"),

    # =========================================================================
    # LEADERSHIP & ROLES — leadershipSkills
    # =========================================================================
    ("Leadership & roles", "Leadership skills object", "—", "Card.leadershipSkills.{commander,brawl,oathbreaker}", "not stored", "faster than iterating legalities for the Commander badge"),
    ("Leadership & roles", "Is Reserved", "Card.reserved (boolean)", "Card.isReserved", "not stored", "could add a 'Reserved List only' filter"),

    # =========================================================================
    # GAMES & PLATFORMS — Card.games[]
    # =========================================================================
    ("Games & platforms", "Games", "Card.games", "Card.games", "filtered to ['paper'] on import", "['paper','mtgo','arena','astral']"),
    ("Games & platforms", "Online-only", "derived from Card.games (includes 'arena' only)", "Card.isOnlineOnly", "not stored", "would filter out Arena-only cards from the binder display"),

    # =========================================================================
    # SETS — full Set object metadata
    # =========================================================================
    ("Sets", "Set UUID", "Set.id", "Set.uuid", "sets.id (uuid)", ""),
    ("Sets", "Set code", "Set.code", "Set.code", "sets.code (text)", ""),
    ("Sets", "Set name", "Set.name", "Set.name", "sets.name (text)", ""),
    ("Sets", "Set release date", "Set.released_at", "Set.releaseDate", "sets.release_date (date)", ""),
    ("Sets", "Set block (we call it series)", "Set.block", "Set.block (rare)", "sets.series (text)", "Scryfall calls it 'block'; we call it 'series'"),
    ("Sets", "Set type", "Set.set_type", "Set.type", "sets.set_type (text)", "30+ enum values"),
    ("Sets", "Card count", "Set.card_count", "Set.totalSetSize", "sets.card_count (int)", ""),
    ("Sets", "Parent set code", "Set.parent_set_code", "Set.parentCode", "not stored", "parent set for child sets"),
    ("Sets", "Digital", "Set.digital", "Set.isOnlineOnly", "filtered out on import (digital sets skipped)", ""),
    ("Sets", "Foil-only", "Set.foil_only", "Set.isFoilOnly", "not stored", "set only available in foil"),
    ("Sets", "Set icon (SVG)", "Set.icon_svg_uri", "Set.svgs.icon", "sets.icon_url (text)", ""),
    ("Sets", "Set search URI", "Set.search_uri", "—", "not stored", "search URL for the set"),
    ("Sets", "Set scry URI", "Set.scry_uri", "—", "not stored", ""),
    ("Sets", "Set URI (Scryfall)", "Set.uri", "—", "not stored", "Scryfall API URI for the set"),
    ("Sets", "Printed size", "Set.printed_size", "— (MTGJSON doesn't expose)", "not stored", "physical set size (vs the larger online-only total)"),
    ("Sets", "TCGplayer set id", "Set.tcgplayer_id", "identifiers.tcgplayerProductId (per card)", "not stored", "TCGplayer set ID"),
    ("Sets", "MTGO code", "Set.mtgo_code", "Set.mtgoCode", "not stored", "MTGO set code (may differ from Set.code)"),
    ("Sets", "Arena code", "Set.arena_code", "Set.arenacode (varies)", "not stored", "Arena set code"),
    ("Sets", "Set legalities (aggregated)", "derived from Card.legalities (aggregated)", "derived from Card.legalities (aggregated)", "sets.legalities (jsonb)", "jsonb_object_agg of legal formats (>50% cards legal) in refreshMtgSetMetadata()"),
    ("Sets", "Set crossover flag", "derived from Card.security_stamp=='triangle' or Card.promo_types?'universesbeyond'", "derived from same Scryfall signals", "sets.crossover (bool)", "ratio of those cards >0.5 = crossover"),
    ("Sets", "Set logo URL", "derived from cards.image_art_crop (highest-priced per set)", "—", "sets.logo_url (text)", "Scryfall has no set key-art — synthesised from highest-priced card"),
    ("Sets", "Purchase URLs", "—", "Card.purchaseUrls.{cardmarket,cardKingdom,tcgplayer}", "not stored", "direct product pages — would let /decks 'buy missing slots' link out per card"),

    # =========================================================================
    # TAXONOMY (APP) — MTG_DECK_TYPES, MTG_BUCKETS, HIDDEN_TYPES, FORMAT_LIST
    # =========================================================================
    ("Taxonomy (app)", "games table", "—", "—", "hardcoded seed: ('mtg', 'Magic: The Gathering'), ('pokemon', 'Pokémon')", ""),
    ("Taxonomy (app)", "ENABLED_GAMES", "—", "—", "hardcoded: ['mtg']", "toggle to surface Pokemon in UI"),
    ("Taxonomy (app)", "MTG_DECK_TYPES", "derived from Set.set_type", "derived from Set.type", "hardcoded: 7 set_type values that are precon products", "excluded from /g/mtg, included on /decks"),
    ("Taxonomy (app)", "MTG_BUCKETS", "—", "—", "hardcoded: 7 buckets mapping set_type → category", "drives /g/mtg tabs and /advisor"),
    ("Taxonomy (app)", "HIDDEN_TYPES (advisor)", "—", "—", "hardcoded: ['token','memorabilia','minigame','vanguard']", "excluded from advisor queries"),
    ("Taxonomy (app)", "FORMAT_LIST (mtg)", "Card.legalities keys", "Card.legalities keys", "hardcoded: ['standard','pioneer','modern','legacy','vintage','commander','pauper']", "drives format radio chips on /g/mtg"),
    ("Taxonomy (app)", "FORMAT_LIST (pokemon)", "—", "—", "hardcoded: ['standard','expanded','unlimited']", "dormant"),
    ("Taxonomy (app)", "POKEMON_AIM", "—", "—", "hardcoded in advisor page: 3 buckets", "dormant"),
    ("Taxonomy (app)", "DEFAULT_AIM (mtg)", "—", "—", "hardcoded: ['core','expansions','crossovers']", "preselected when ?aim is absent"),

    # =========================================================================
    # FACETS (DERIVED) — card_facets rows
    # =========================================================================
    ("Facets (derived)", "color facet", "derived from Card.colors (WUBRG) or Card.card_faces[].colors", "Card.colors", "card_facets rows", "derived: uniq colors sorted WUBRG, or 'C' for colourless (mtgFacetsFor)"),
    ("Facets (derived)", "color_combo facet", "derived from Card.colors or Card.card_faces[].colors", "Card.colors", "card_facets rows", "derived: sorted WUBRG letters joined, or 'C' (mtgFacetsFor)"),
    ("Facets (derived)", "kind facet", "derived from Card.type_line or Card.card_faces[].type_line", "Card.type (or Card.types[])", "card_facets rows", "derived: MTG_KINDS.find(type_line) — hand-rolled kind ordering"),

    # =========================================================================
    # CONSTANTS (APP) — WUBRG, MTG_KINDS, MTG_RARITY_TIER, sortKey
    # =========================================================================
    ("Constants (app)", "WUBRG order", "—", "—", "hardcoded 'WUBRG' in util.ts and ComboSlicer.tsx", "standard MTG order — neither source enumerates it"),
    ("Constants (app)", "MTG_KINDS list", "—", "—", "hardcoded in util.ts", "matches Scryfall type_line tokens"),
    ("Constants (app)", "MTG_RARITY_TIER map", "—", "—", "hardcoded 1-5 map in util.ts", "neither source provides a tier; we map from rarity text"),
    ("Constants (app)", "sortKey() heuristic", "—", "—", "hand-rolled regex/sort", ""),

    # =========================================================================
    # UX KNOBS (APP) — defaults & thresholds
    # =========================================================================
    ("UX knobs (app)", "min set size default", "—", "—", "hardcoded: 10", "tiny promo sets filter out"),
    ("UX knobs (app)", "binder cover marquee rule", "—", "—", "derived: highest-priced card's art_crop per set — DISTINCT ON (set_id) ORDER BY price.usd DESC", ""),
    ("UX knobs (app)", "download() cache TTL", "—", "—", "hardcoded: 20h", "skip re-download if file <20h old"),
    ("UX knobs (app)", "Batch sizes (bulk import)", "—", "—", "hardcoded: 1000 card-batch / 5000 facet+price batches", ""),

    # =========================================================================
    # NOT IMPORTED — caught explicitly to remind ourselves
    # =========================================================================
    ("Not imported", "Rulings", "GET /cards/{id}/rulings (Card.rulings_uri)", "Card.rulings[].date / Card.rulings[].text (bundled)", "not imported", "per-card fetch adds ~25MB; never used"),
    ("Not imported", "Scryfall API URIs", "Card.{uri,scryfall_uri,set_uri,set_search_uri,scryfall_set_uri,rulings_uri,prints_search_uri}", "—", "not stored", "Scryfall API/web URLs — debug-only"),
]


def write_app_sheet(ws, rows):
    """Single sheet: one row per conceptual element, one column per source.

    Columns: Group | Element | Scryfall | MTGJSON | App today | # Scryfall | # MTGJSON | Status | Why
      Each source column contains the actual schema field path (e.g. 'Card.id',
      'Card.image_uris.art_crop', 'Card.identifiers.scryfallId', 'Set.code'),
      or 'derived' / '—' / 'hardcoded' to describe the source shape.
      The two `#` columns show how many records each field would populate across
      the full bulk download — written from docs/schema-counts.json if present.
      Status is the annotation column for the final decision.
    """
    counts = {}
    try:
        with open(COUNTS_PATH, encoding="utf-8") as f:
            counts = json.load(f).get("counts", {})
    except FileNotFoundError:
        pass

    headers = ("Group", "Element", "Scryfall field / source", "MTGJSON field / source",
               "App today (current)", "# Scryfall", "# MTGJSON", "Status", "Why")
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 36
    for group, element, scryfall, mtgjson, app, why in rows:
        key = f"{group}::{element}"
        s_count = counts.get(key, {}).get("scryfall")
        m_count = counts.get(key, {}).get("mtgjson")
        ws.append((group, element, scryfall, mtgjson, app, s_count, m_count, None, why))
    widths = [22, 46, 38, 38, 38, 12, 12, 14, 80]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(bold=True, italic=True, color="4B5563")
        for c in range(2, 6):
            ws.cell(row=r, column=c).alignment = WRAP
        for c in (6, 7, 8):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=9).alignment = WRAP
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"


def main():
    # Write a JSON sidecar with the same row order as the xlsx so
    # scripts/count_schema_fields.ts can read it without an xlsx parser dep.
    with open(ROWS_PATH, "w", encoding="utf-8") as f:
        json.dump(
            [{"group": g, "element": e, "scryfall": s, "mtgjson": m}
             for (g, e, s, m, _app, _why) in APP_ELEMENTS],
            f, indent=2,
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sources"
    write_app_sheet(ws, APP_ELEMENTS)

    ws_legend = wb.create_sheet("Legend")
    ws_legend.append(["Column", "Meaning"])
    for cell in ws_legend[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    legend = [
        ("Group", "Anatomical bucket (Identity / Card text / Picture / Gameplay / Faces / Colors / Types / Frame & finish / Print / Promo / Identifiers / Prices / Legalities / Related cards / Leadership / Games / Sets / Taxonomy / Facets / Constants / UX knobs / Not imported)"),
        ("Element", "The semantic thing this row represents (a column, a derived value, a constant)"),
        ("Scryfall field / source", "The actual Scryfall Card/Set field path, or 'derived' / '—' if the source has nothing for it"),
        ("MTGJSON field / source", "The actual MTGJSON Card/Set field path, or 'derived' / '—' if the source has nothing for it"),
        ("App today (current)", "What we currently store/derive/use in this app, or 'not used' / 'not stored'"),
        ("# Scryfall", "How many records would populate this field across the full Scryfall default_cards bulk — populated by `npm run schema:counts`"),
        ("# MTGJSON", "How many records would populate this field across the full MTGJSON AllPrintings.json bulk — populated by `npm run schema:counts`"),
        ("Status", "Blank — annotate with yes / no / ? to make the decision"),
        ("Why", "Notes for the Status decision"),
    ]
    for r, row in enumerate(legend, start=2):
        ws_legend.append(row)
        ws_legend.cell(row=r, column=1).font = Font(bold=True)
        for c in range(1, 3):
            ws_legend.cell(row=r, column=c).alignment = WRAP
    ws_legend.column_dimensions["A"].width = 22
    ws_legend.column_dimensions["B"].width = 110

    wb.save("docs/schema-scryfall-vs-mtgjson.xlsx")
    print(f"wrote docs/schema-scryfall-vs-mtgjson.xlsx ({len(APP_ELEMENTS)} rows)")
    print(f"wrote {ROWS_PATH} ({len(APP_ELEMENTS)} rows)")


if __name__ == "__main__":
    main()
